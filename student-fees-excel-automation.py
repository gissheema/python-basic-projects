import pandas as pd # pandas
from openpyxl import load_workbook # openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# =========================
# CONFIG
# =========================
INPUT_FILE = "students.xlsx"
OUTPUT_FILE = "fees_report.xlsx"
LATE_FEE_PER_DAY = 50

# =========================
# LOAD DATA
# =========================
df = pd.read_excel(INPUT_FILE)

# Convert DueDate to datetime
df["DueDate"] = pd.to_datetime(df["DueDate"])

today = datetime.today()

# =========================
# CALCULATIONS
# =========================

# Total Fees
df["Total Fees"] = df["Tuition"] + df["Transport"] + df["Hostel"]

# Late Fee Calculation
def calculate_late_fee(row):
    if row["Paid"] >= row["Total Fees"]:
        return 0
    delay = (today - row["DueDate"]).days
    return max(0, delay * LATE_FEE_PER_DAY)

df["Late Fee"] = df.apply(calculate_late_fee, axis=1)

# Final Amount
df["Final Amount"] = df["Total Fees"] + df["Late Fee"]

# Balance
df["Balance"] = df["Final Amount"] - df["Paid"]

# Status
df["Status"] = df["Balance"].apply(
    lambda x: "Paid" if x <= 0 else "Pending"
)

# =========================
# CLASS SUMMARY
# =========================
summary = df.groupby("Class").agg({
    "Final Amount": "sum",
    "Paid": "sum",
    "Balance": "sum"
}).reset_index()

# =========================
# SAVE TO EXCEL
# =========================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Student Report", index=False)
    summary.to_excel(writer, sheet_name="Class Summary", index=False)

# =========================
# FORMATTING
# =========================
wb = load_workbook(OUTPUT_FILE)

# Sheets
ws1 = wb["Student Report"]
ws2 = wb["Class Summary"]

# Styles
header_font = Font(bold=True)
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# Format headers
for cell in ws1[1]:
    cell.font = header_font

for cell in ws2[1]:
    cell.font = header_font

# Highlight rows based on status
for row in ws1.iter_rows(min_row=2):
    status = row[-1].value
    if status == "Pending":
        for cell in row:
            cell.fill = red_fill
    else:
        for cell in row:
            cell.fill = green_fill

# Auto column width
def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

auto_width(ws1)
auto_width(ws2)

# Save workbook
wb.save(OUTPUT_FILE)

print("✅ Full fees automation report generated: fees_report.xlsx")